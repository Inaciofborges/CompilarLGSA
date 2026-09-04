#!/usr/bin/env python3
"""
Script para compilar dados de poços a partir de arquivos Excel.

Extrai informações de células específicas e gera um arquivo de saída
com a estrutura:
  Well | MD | Size | Volume
       |    | mm   | %
"""

import os
import sys
from pathlib import Path
from openpyxl import load_workbook
import csv

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

# Tabela de classificação de tamanho de partículas (escala de Wentworth)
SIZE_CLASSES = [
    (2.0, "Granule"),
    (1.682, "Very Coarse Sand"),
    (0.841, "Coarse Sand"),
    (0.42, "Medium Sand"),
    (0.21, "Fine Sand"),
    (0.105, "Very Fine Sand"),
    (0.053, "Coarse Silt"),
    (0.026, "Medium Silt"),
    (0.013, "Fine Silt"),
    (0.007, "Very Fine Silt"),
    (0.0, "Clay")
]

def classify_size(size_value):
    """
    Classifica o tamanho da partícula baseado na escala de Wentworth.

    Args:
        size_value: Valor do tamanho em mm

    Returns:
        str: Classificação do tamanho (ex: "Medium Sand")
    """
    if size_value is None:
        return ""

    for threshold, class_name in SIZE_CLASSES:
        if size_value >= threshold:
            return class_name

    return "Clay"


def extract_well_data(file_path):
    """
    Extrai dados de um arquivo Excel (.xlsx ou .xls).

    Args:
        file_path (str): Caminho do arquivo Excel

    Returns:
        list: Lista de dicionários com os dados extraídos (um por linha de amostra)
    """
    try:
        file_ext = Path(file_path).suffix.lower()

        if file_ext == '.xls':
            # Lê arquivo .xls usando xlrd
            if not HAS_XLRD:
                print(f"Erro: xlrd não está instalado. Use: pip install xlrd --user", file=sys.stderr)
                return None

            import xlrd
            wb = xlrd.open_workbook(file_path)
            ws = wb.sheet_by_index(0)

            # Extrai Well, Amostra e MD (xlrd usa 0-indexed)
            well = ws.cell_value(2, 0)    # A3
            amostra = ws.cell_value(2, 14)  # O3
            md = ws.cell_value(3, 14)    # O4

            # Extrai Size (A75:A85) e Volume (F75:F85)
            size_values = []
            volume_values = []
            for row_idx in range(74, 85):  # Linhas 75-85 (0-indexed: 74-84)
                try:
                    size_val = ws.cell_value(row_idx, 0)  # Coluna A
                    if size_val != '':  # Verifica se não é string vazia
                        size_values.append(float(size_val) if size_val else None)
                    else:
                        size_values.append(None)
                except (ValueError, TypeError):
                    size_values.append(None)

                try:
                    vol_val = ws.cell_value(row_idx, 5)  # Coluna F
                    if vol_val != '':  # Verifica se não é string vazia
                        volume_values.append(float(vol_val) if vol_val else None)
                    else:
                        volume_values.append(None)
                except (ValueError, TypeError):
                    volume_values.append(None)

        else:
            # Lê arquivo .xlsx usando openpyxl
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            # Extrai Well, Amostra e MD
            well = ws['A3'].value
            amostra = ws['O3'].value
            md = ws['O4'].value

            # Extrai Size (A75:A85) e Volume (F75:F85)
            size_values = []
            volume_values = []
            for row in range(75, 86):  # Linhas 75-85
                size_val = ws[f'A{row}'].value
                size_values.append(float(size_val) if size_val else None)

                vol_val = ws[f'F{row}'].value
                volume_values.append(float(vol_val) if vol_val else None)

        # Debug: Mostra o que foi extraído
        print(f"    Well: {well}, MD: {md}, Amostra: {amostra}")
        print(f"    Size values ({len(size_values)}): {size_values[:3]}...")
        print(f"    Volume values ({len(volume_values)}): {volume_values[:3]}...")

        # Cria lista de dados com uma linha para cada amostra
        data_list = []
        for i in range(max(len(size_values), len(volume_values))):
            size_val = size_values[i] if i < len(size_values) else None
            data_list.append({
                'Well': well,
                'MD': md,
                'Amostra': amostra,
                'Size Class': classify_size(size_val),
                'Size': size_val,
                'Volume': volume_values[i] if i < len(volume_values) else None
            })

        return data_list

    except Exception as e:
        print(f"Erro ao processar {file_path}: {e}", file=sys.stderr)
        return None


def compile_well_data(input_folder, output_file=None):
    """
    Processa todos os arquivos Excel de uma pasta e compila os dados.

    Args:
        input_folder (str): Pasta contendo os arquivos Excel
        output_file (str): Arquivo de saída (CSV)
    """
    input_path = Path(input_folder)

    if not input_path.exists():
        print(f"Erro: A pasta '{input_folder}' não existe.", file=sys.stderr)
        return False

    # Encontra todos os arquivos Excel
    excel_files = list(input_path.glob('*.xlsx')) + list(input_path.glob('*.xls'))

    if not excel_files:
        print(f"Aviso: Nenhum arquivo Excel encontrado em '{input_folder}'", file=sys.stderr)
        return False

    compiled_data = []
    well_name = None

    print(f"Processando {len(excel_files)} arquivo(s)...")
    for excel_file in excel_files:
        print(f"  Processando: {excel_file.name}")
        well_data = extract_well_data(str(excel_file))
        if well_data:
            compiled_data.extend(well_data)
            # Extrai o nome do poço da primeira entrada
            if well_name is None and well_data[0].get('Well'):
                well_name = well_data[0].get('Well')

    if not compiled_data:
        print("Nenhum dado foi extraído.", file=sys.stderr)
        return False

    # Gera o nome do arquivo se não foi fornecido
    if output_file is None:
        if well_name:
            output_file = f"{well_name}_LGSA.csv"
        else:
            output_file = 'compiled_wells.csv'
            print("Aviso: Não foi possível determinar o nome do poço. Usando nome padrão.", file=sys.stderr)

    # Escreve no arquivo de saída
    output_path = Path(output_file)

    try:
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter='\t')

            # Primeira linha: nomes das colunas (sem unidades)
            writer.writerow(['Well', 'MD', 'Amostra', 'Size Class', 'Size', 'Volume'])

            # Segunda linha: unidades
            writer.writerow(['', '', '', '', 'mm', '%'])

            # Dados compilados
            for data in compiled_data:
                size_val = data.get('Size')
                volume_val = data.get('Volume')

                writer.writerow([
                    data.get('Well', ''),
                    data.get('MD', ''),
                    data.get('Amostra', ''),
                    data.get('Size Class', ''),
                    size_val if size_val is not None else '',
                    volume_val if volume_val is not None else ''
                ])

        print(f"\n✓ Dados compilados com sucesso em: {output_path.absolute()}")
        return True

    except Exception as e:
        print(f"Erro ao escrever arquivo de saída: {e}", file=sys.stderr)
        return False


if __name__ == '__main__':
    print("=" * 60)
    print("Script de Compilação de Dados de Poços")
    print("=" * 60)

    # Se argumentos foram passados, usa modo linha de comando
    if len(sys.argv) >= 2:
        input_folder = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else None
    else:
        # Modo interativo
        print("\nDigite o caminho da pasta contendo os arquivos Excel:")
        print("(Exemplo: ./dados ou C:\\Users\\seu_usuario\\dados)")
        input_folder = input("\nCaminho da pasta: ").strip()

        if not input_folder:
            print("\nErro: Caminho não pode estar vazio!")
            sys.exit(1)

        output_file = None
        print("\n" + "=" * 60)

    success = compile_well_data(input_folder, output_file)

    if success:
        print("\n" + "=" * 60)
        print("✓ Processamento concluído com sucesso!")
        print("=" * 60)

    sys.exit(0 if success else 1)
