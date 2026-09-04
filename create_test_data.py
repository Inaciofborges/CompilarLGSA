#!/usr/bin/env python3
"""
Script para criar arquivos Excel de teste.
Útil para testar o script compile_well_data.py
"""

from openpyxl import Workbook
from pathlib import Path
import sys


def create_test_excel(filename, well_name, md_value, amostra_value):
    """
    Cria um arquivo Excel de teste com dados nas células específicas.

    Args:
        filename (str): Nome do arquivo a criar
        well_name (str): Nome do poço
        md_value (float): Valor de MD
        amostra_value (str): Nome da amostra
    """
    wb = Workbook()
    ws = wb.active

    # Adiciona dados nas células especificadas
    ws['A3'] = well_name      # Well
    ws['O4'] = md_value       # MD
    ws['O3'] = amostra_value  # Amostra

    # Adiciona valores de Size (A75:A85)
    size_values = [2.1, 2.3, 2.2, 2.4, 2.2, 2.5, 2.3, 2.1, 2.4, 2.2, 2.3]
    for idx, val in enumerate(size_values):
        ws[f'A{75 + idx}'] = val

    # Adiciona valores de Volume (F75:F85)
    volume_values = [65.5, 67.2, 66.8, 68.1, 67.0, 69.5, 68.3, 66.5, 68.8, 67.1, 68.0]
    for idx, val in enumerate(volume_values):
        ws[f'F{75 + idx}'] = val

    wb.save(filename)
    print(f"✓ Arquivo criado: {filename}")


def main():
    """Cria arquivos de teste na pasta test_data"""
    test_dir = Path('./test_data')
    test_dir.mkdir(exist_ok=True)

    print("Criando arquivos de teste...\n")

    # Cria alguns arquivos de exemplo
    test_files = [
        ('WELL-001.xlsx', 'WELL-001', 1500.0, 'Sample-A1'),
        ('WELL-002.xlsx', 'WELL-002', 1650.5, 'Sample-B1'),
        ('WELL-003.xlsx', 'WELL-003', 1725.0, 'Sample-C1'),
    ]

    for filename, well, md, amostra in test_files:
        filepath = test_dir / filename
        create_test_excel(str(filepath), well, md, amostra)

    print(f"\n✓ Arquivos de teste criados em: {test_dir.absolute()}")
    print(f"\nAgora você pode testar o script com:")
    print(f"  python compile_well_data.py ./test_data")


if __name__ == '__main__':
    main()
